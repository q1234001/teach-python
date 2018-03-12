import openpyxl

#宣告一個試算表
workbook = openpyxl.Workbook()
workbook.remove_sheet(workbook.get_sheet_by_name('Sheet'))
workbook.create_sheet('編號1')
workbook.create_sheet('編號2')
workbook.create_sheet(index = 1, title = '編號之我插隊')


#操作一個工作表
sheet = workbook.get_sheet_by_name('編號1')

#寫入值
sheet['A1'] = 'ABC'
#是從'1'開始
sheet.cell(row = 1, column = 2).value = 'CDE'
#記得存檔歐
workbook.save('test1.xlsx')

