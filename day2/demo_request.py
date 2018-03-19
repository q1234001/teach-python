import json
import requests
import openpyxl

url = 'https://exhibitorsearch.messefrankfurt.com/service/esb/1.0/search/exhibitor?language=en&q=&orderBy=stand&pageNumber=1&pageSize=1632&showJumpLabels=true&jumpLabelId=&eventId=PAPERWORLD'
get_data = requests.get(url)
workbook = openpyxl.Workbook()
sheet = workbook.get_sheet_by_name('Sheet')
sheet['A1'] = 'No.'
sheet['B1'] = 'Exhibitor  (參展商)'
sheet['C1'] = 'Booth no.  (攤位號碼)'
sheet['D1'] = 'Contact person (聯絡人)'
sheet['E1'] = 'E-mail / web site (EMAIL 或是 網站)'
sheet['F1'] = 'Remarks  (備註)' 







print (get_data.json()['result']['hits'][0]['exhibitor']['name'])
print (get_data.json()['result']['hits'][0]['exhibitor']["exhibition"]['exhibitionHall'][0]['name'])
print (get_data.json()['result']['hits'][0]['exhibitor']["exhibition"]['exhibitionHall'][0]['stand'][0]['name'])
print (get_data.json()['result']['hits'][0]['exhibitor']['address']['tel'] )
print (get_data.json()['result']['hits'][0]['exhibitor']['address']['fax'] )
print (get_data.json()['result']['hits'][0]['exhibitor']['homepage'])
print (get_data.json()['result']['hits'][0]['exhibitor']['categories'][1]['subCategories'][0]["subCategories"][0]['name'])






x = 0



while x < len(get_data.json()['result']['hits']):
    
    sheet.cell(row = x+2 ,column = 1).value = x+1
    sheet.cell(row = x+2 ,column = 2).value = get_data.json()['result']['hits'][x]['exhibitor']['name']
    sheet.cell(row = x+2 ,column = 3).value = get_data.json()['result']['hits'][x]['exhibitor']["exhibition"]['exhibitionHall'][0]['name'] + " " + get_data.json()['result']['hits'][x]['exhibitor']["exhibition"]['exhibitionHall'][0]['stand'][0]['name']

    if get_data.json()['result']['hits'][x]['exhibitor']['address']['tel'] != None:
        sheet.cell(row = x+2 ,column = 4).value = get_data.json()['result']['hits'][x]['exhibitor']['address']['tel']
    elif get_data.json()['result']['hits'][x]['exhibitor']['address']['fax'] != None:
        sheet.cell(row = x+2 ,column = 4).value = get_data.json()['result']['hits'][x]['exhibitor']['address']['fax']
    else:
        sheet.cell(row = x+2 ,column = 4).value = 'n/a'

    if get_data.json()['result']['hits'][0]['exhibitor']['address']['email'] != None:
         sheet.cell(row = x+2 ,column = 5).value = get_data.json()['result']['hits'][x]['exhibitor']['address']['email']
    elif get_data.json()['result']['hits'][0]['exhibitor']['homepage'] != None:
        sheet.cell(row = x+2 ,column = 5).valueget_data.json()['result']['hits'][x]['exhibitor']['homepage']
    else:
        sheet.cell(row = x+2 ,column = 5).value = 'n/a'
    
    y = 0
    pro =''

    while y < len(get_data.json()['result']['hits'][x]['exhibitor']['categories']):
        try:
            try:
                sheet.cell(row = x+2 ,column = 6).value = str(sheet.cell(row = x+2 ,column = 6).value) + "/" + get_data.json()['result']['hits'][x]['exhibitor']['categories'][y]['subCategories'][0]["subCategories"][0]['name']
            except TypeError:
                sheet.cell(row = x+2 ,column = 6).value = str(sheet.cell(row = x+2 ,column = 6).value) + "/" + get_data.json()['result']['hits'][x]['exhibitor']['categories'][y]['subCategories'][0]['name']
        except TypeError:
            sheet.cell(row = x+2 ,column = 6).value = 'error'
                    
                    
                     
                
                                
        y+=1
        workbook.save('test2.xlsx')  
        
    now = str(x+1)+ "/" +str(len(get_data.json()['result']['hits']))
    print (now + " down")
    x+=1

    




